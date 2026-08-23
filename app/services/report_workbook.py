"""Turns a report payload into a formatted .xlsx.

One builder for every report. It knows about spreadsheets — fonts, widths,
freeze panes, number formats — and nothing about farmers, stock or baskets. The
report services describe their own shape in the payload they return, so adding a
third report means writing a query, not another copy of this file.

Called from four places: the cron publish job, the admin Export buttons on the
website and in the app, and `tools/f2h_report.py` run by hand. Deliberately free
of Flask and SQLAlchemy imports, because only some of those have an application
context.

── The payload ────────────────────────────────────────────────────────────────

    {
      "title":       "F2H Market — farmers, products and stock",
      "sheet_name":  "Farmers & Stock",
      "subtitle":    "optional line under the title",
      "generated_at": "2026-08-22T04:30:00Z",
      "columns": [
          {"key": "farm_name", "label": "Farm", "width": 24},
          {"key": "price", "label": "Price", "format": "money"},
      ],
      "rows": [ {...}, ... ],          # one dict per row, keyed by column key
      "summary": {...},                # headline counts, for the caller's logs
      "summary_rows": [                # optional Summary tab
          {"label": "Products out of stock",
           "formula": "COUNTIF({available_quantity},0)"}
      ],
    }

`{column_key}` inside a formula is replaced with that column's absolute range on
the data sheet. Writing `COUNTIF({available_quantity},0)` rather than
`COUNTIF('Farmers & Stock'!$J$5:$J$88,0)` keeps the formula readable and means a
column moving left or right cannot silently point it at the wrong data.
"""

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Arial throughout, per house style for anything that leaves the building.
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=14, bold=True)
NOTE = Font(name="Arial", size=9, italic=True, color="666666")

HEAD_FILL = PatternFill("solid", fgColor="2F6B2F")

# Rows a report wants to draw attention to. The report decides *which* rows by
# setting `_highlight`; this file decides what that looks like, so the two
# reports cannot end up with different shades of "needs attention".
HIGHLIGHTS = {
    "warn": PatternFill("solid", fgColor="FDE7E9"),   # wrong, or empty
    "note": PatternFill("solid", fgColor="FFF6DA"),   # unusual, worth a look
    "info": PatternFill("solid", fgColor="E8F1FB"),   # not yet real
}

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Named formats, so a report says what a column *is* rather than repeating an
# Excel format string and getting a decimal place different somewhere.
FORMATS = {
    "money": "₹#,##0.00",
    "quantity": "#,##0.###",
    "integer": "#,##0",
    "date": "yyyy-mm-dd",
}

HEADER_ROW = 4


def build(payload, path):
    """Write the workbook to `path` (a filename or a file-like object)."""
    columns = payload["columns"]
    rows = payload["rows"]
    keys = [c["key"] for c in columns]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = payload.get("sheet_name", "Report")

    generated = payload.get("generated_at", "")
    sheet["A1"] = payload.get("title", "F2H Market report")
    sheet["A1"].font = TITLE
    sheet["A2"] = payload.get(
        "subtitle",
        f"Generated {generated} (UTC) from the live database. "
        "Do not edit — this file is overwritten on each run.")
    sheet["A2"].font = NOTE

    for i, column in enumerate(columns, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=i, value=column["label"])
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER

    formats = {c["key"]: FORMATS.get(c.get("format")) for c in columns}

    for r, row in enumerate(rows, start=HEADER_ROW + 1):
        fill = HIGHLIGHTS.get(row.get("_highlight"))
        for c, key in enumerate(keys, start=1):
            cell = sheet.cell(row=r, column=c, value=row.get(key))
            cell.font = BODY
            cell.border = BORDER
            if formats.get(key) and row.get(key) is not None:
                cell.number_format = formats[key]
            if fill:
                cell.fill = fill

    last = HEADER_ROW + len(rows)

    # Freeze under the header and turn on filtering: these are sheets people
    # sort and filter, and both are one click from here.
    sheet.freeze_panes = sheet.cell(row=HEADER_ROW + 1, column=1)
    if rows:
        sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(keys))}{last}"

    for i, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = column.get("width", 14)
    sheet.row_dimensions[HEADER_ROW].height = 30

    if payload.get("summary_rows"):
        _summary(workbook, payload, keys, last)

    workbook.save(path)
    return len(rows)


def _summary(workbook, payload, keys, last):
    """A Summary tab whose figures are formulas over the data sheet.

    Formulas rather than constants on purpose: someone will filter or delete
    rows in their own copy, and a hardcoded total would then contradict the
    sheet sitting next to it. Everything used is Excel-2007-era, so LibreOffice
    evaluates it too — no `_xlfn.` prefix and nothing that spills.
    """
    sheet = workbook.create_sheet("Summary")
    data_sheet = payload.get("sheet_name", "Report")
    first = HEADER_ROW + 1

    def ranges(formula):
        """Replace every `{column_key}` with that column's absolute range."""
        def replace(match):
            key = match.group(1)
            if key not in keys:
                raise KeyError(
                    f"Summary formula references unknown column {key!r}. "
                    f"Known columns: {', '.join(keys)}")
            letter = get_column_letter(keys.index(key) + 1)
            return f"'{data_sheet}'!${letter}${first}:${letter}${last}"
        return re.sub(r"\{(\w+)\}", replace, formula)

    sheet["A1"] = "Summary"
    sheet["A1"].font = TITLE
    sheet["A2"] = f"Generated {payload.get('generated_at', '')} (UTC)"
    sheet["A2"].font = NOTE

    for i, entry in enumerate(payload["summary_rows"], start=4):
        sheet.cell(row=i, column=1, value=entry["label"]).font = BOLD
        cell = sheet.cell(row=i, column=2, value="=" + ranges(entry["formula"]))
        cell.font = BODY
        cell.number_format = FORMATS.get(entry.get("format"), "#,##0.###")

    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 16

    note = sheet.cell(row=len(payload["summary_rows"]) + 6, column=1)
    note.value = (f"Counts are formulas over the '{data_sheet}' tab, so they "
                  "follow the data if rows are added or removed.")
    note.font = NOTE


def build_bytes(payload):
    """The workbook as bytes, for sending straight down an HTTP response.

    Saving to a real file first and reading it back would need somewhere
    writable, which a container may not have, and leaves a file to clean up on
    every download.
    """
    buffer = io.BytesIO()
    build(payload, buffer)
    return buffer.getvalue()


def filename(payload):
    """A stable filename — `F2H-farmer-stock.xlsx`, with no date in it.

    The date used to be part of the name, which quietly defeated the whole
    point of `drive_upload.upload()`: it finds an existing file *by name* and
    updates it in place, so a name that changes at midnight meant a brand new
    file on every run after the first of each day. Three reports at their
    current cadences would have left roughly seven hundred near-identical
    spreadsheets in one folder over a year, and the Drive link anybody had
    bookmarked would have pointed at a stale one within days.

    Undated, each report is one file that is always current, at one URL. The
    generation timestamp has not been lost — it is on the sheet itself, under
    the title and again on the Summary tab, which is where somebody reading the
    file will look for it rather than at the filename.

    History is not lost either: updating in place is exactly what makes Drive
    keep the previous contents in its own revision history, where they can be
    restored without cluttering the folder.
    """
    slug = payload.get("slug", "report")
    return f"F2H-{slug}.xlsx"


def download_filename(payload):
    """A dated name — `F2H-farmer-stock-2026-08-22.xlsx` — for a copy leaving.

    The opposite choice to [filename], and for the opposite reason. The Drive
    copy is a living document that is overwritten in place, so a stable name is
    what keeps its link working. A download or a share is a *snapshot* somebody
    is taking out of the system: it lands in a Downloads folder or a chat
    beside other files, will be looked at weeks later, and the one thing its
    reader will want to know is which day it describes.
    """
    stamp = (payload.get("generated_at") or "")[:10]
    slug = payload.get("slug", "report")
    return f"F2H-{slug}-{stamp}.xlsx" if stamp else f"F2H-{slug}.xlsx"
